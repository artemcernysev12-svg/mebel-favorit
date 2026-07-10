#!/usr/bin/env python3
"""Download the latest stock workbook from Yandex Mail and install it safely."""

from __future__ import annotations

import hashlib
import imaplib
import os
import re
import ssl
import stat
import sys
import tempfile
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from email import policy
from email.header import decode_header
from email.message import Message
from email.parser import BytesParser
from email.utils import getaddresses
from enum import Enum
from itertools import islice
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # Handled explicitly before an attachment can be installed.
    load_workbook = None


IMAP_HOST = "imap.yandex.com"
IMAP_PORT = 993
MAILBOX = "INBOX"
# Адрес поставщика берётся из GitHub Secret STOCK_SENDER_EMAIL, чтобы не
# раскрывать его в публичном репозитории. Значение по умолчанию — только
# заглушка для офлайн-тестов; в рабочем запуске обязателен реальный секрет.
EXPECTED_SENDER = os.environ.get("STOCK_SENDER_EMAIL", "supplier@example.com").strip()
SUBJECT_FRAGMENT = "Остатки товаров"
DEFAULT_MAX_AGE_HOURS = 72
TARGET_RELATIVE_PATH = Path("stock-data") / "stock.xlsx"

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
REQUIRED_XLSX_MEMBERS = frozenset(
    {"[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"}
)
MAX_ATTACHMENT_BYTES = 95 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
MAX_SUSPICIOUS_COMPRESSION_RATIO = 200
MAX_OPENPYXL_ROWS_TO_SCAN = 50_000

_MONTH_NAMES = (
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
)
_MONTH_NUMBERS = {name.casefold(): index for index, name in enumerate(_MONTH_NAMES, 1)}
_INTERNALDATE_RE = re.compile(
    rb'INTERNALDATE\s+"( ?\d{1,2})-([A-Za-z]{3})-(\d{4}) '
    rb'(\d{2}):(\d{2}):(\d{2}) ([+-])(\d{2})(\d{2})"',
    re.IGNORECASE,
)


class StockUpdateError(Exception):
    """Base class for expected, sanitized failures."""


class ConfigurationError(StockUpdateError):
    """The workflow or local environment is configured incorrectly."""


class MailboxError(StockUpdateError):
    """The IMAP connection or protocol exchange failed."""


class AttachmentValidationError(StockUpdateError):
    """A selected attachment is not a safe, valid XLSX workbook."""


class InstallError(StockUpdateError):
    """A validated workbook could not be installed atomically."""


class Outcome(str, Enum):
    UPDATED = "updated"
    UNCHANGED = "unchanged"
    NO_MESSAGE = "no_message"
    NO_ATTACHMENT = "no_attachment"


@dataclass(frozen=True)
class UpdateResult:
    outcome: Outcome
    changed: bool


@dataclass(frozen=True)
class AttachmentCandidate:
    part: Message
    filename: str
    index: int
    score: int


def _normalized(value: str) -> str:
    return unicodedata.normalize("NFKC", value)


def decode_mime_header(value: Any) -> str:
    """Decode RFC 2047/RFC 2231 text, including common legacy Cyrillic."""
    if value is None:
        return ""

    raw_value = str(value)
    try:
        fragments = decode_header(raw_value)
    except (LookupError, ValueError):
        return _normalized(raw_value)

    decoded: list[str] = []
    for fragment, charset in fragments:
        if isinstance(fragment, str):
            decoded.append(fragment)
            continue

        encodings = [charset] if charset else []
        encodings.extend(["utf-8", "cp1251", "latin-1"])
        for encoding in encodings:
            if not encoding:
                continue
            try:
                decoded.append(fragment.decode(encoding))
                break
            except (LookupError, UnicodeDecodeError):
                continue
        else:
            decoded.append(fragment.decode("utf-8", errors="replace"))

    return _normalized("".join(decoded))


def safe_log_filename(filename: str) -> str:
    """Prevent an untrusted MIME filename from forging workflow log lines."""
    cleaned = "".join(
        " " if unicodedata.category(character).startswith("C") else character
        for character in filename
    )
    cleaned = " ".join(cleaned.split())
    return (cleaned or "вложение без имени")[:160]


def _filename_basename(filename: str) -> str:
    return filename.replace("\\", "/").rsplit("/", 1)[-1].strip()


def format_imap_date(value: date) -> str:
    """Format an IMAP date without depending on the runner locale."""
    return f"{value.day:02d}-{_MONTH_NAMES[value.month - 1]}-{value.year:04d}"


def parse_internaldate(metadata: bytes) -> datetime:
    """Return a message's server-side INTERNALDATE in UTC."""
    match = _INTERNALDATE_RE.search(metadata)
    if not match:
        raise MailboxError("Яндекс Почта не вернула дату получения найденного письма.")

    day, month_name, year, hour, minute, second, sign, offset_hour, offset_minute = (
        match.groups()
    )
    month = _MONTH_NUMBERS.get(month_name.decode("ascii").casefold())
    if month is None:
        raise MailboxError("Яндекс Почта вернула некорректную дату получения письма.")

    offset = timedelta(hours=int(offset_hour), minutes=int(offset_minute))
    if sign == b"-":
        offset = -offset

    try:
        value = datetime(
            int(year),
            month,
            int(day),
            int(hour),
            int(minute),
            int(second),
            tzinfo=timezone(offset),
        )
    except ValueError as error:
        raise MailboxError("Яндекс Почта вернула некорректную дату получения письма.") from error
    return value.astimezone(timezone.utc)


def _status_is_ok(status: Any) -> bool:
    if isinstance(status, bytes):
        status = status.decode("ascii", errors="ignore")
    return str(status).upper() == "OK"


def _extract_fetch_response(response: Any) -> tuple[bytes, bytes]:
    if not response:
        raise MailboxError("Яндекс Почта вернула пустой ответ при чтении письма.")

    for item in response:
        if not isinstance(item, tuple) or len(item) < 2:
            continue
        metadata, payload = item[0], item[1]
        if isinstance(metadata, bytes) and isinstance(payload, bytes):
            return metadata, payload

    raise MailboxError("Не удалось разобрать ответ Яндекс Почты при чтении письма.")


def _internaldate_fragments(response: Any) -> bytes:
    """Join the non-literal parts of a FETCH response for INTERNALDATE search.

    When one FETCH mixes INTERNALDATE with a BODY literal, some servers (Yandex
    among them) return INTERNALDATE in a trailing fragment after the literal
    rather than in the tuple prefix. Scanning every non-literal fragment finds
    it regardless of ordering, without ever scanning the header payload itself.
    """
    fragments: list[bytes] = []
    for item in response or []:
        if isinstance(item, tuple) and item and isinstance(item[0], (bytes, bytearray)):
            fragments.append(bytes(item[0]))
        elif isinstance(item, (bytes, bytearray)):
            fragments.append(bytes(item))
    return b" ".join(fragments)


def _message_matches(message: Message) -> bool:
    sender_headers = [decode_mime_header(value) for value in message.get_all("From", [])]
    addresses = {
        address.casefold()
        for _display_name, address in getaddresses(sender_headers)
        if address
    }
    if EXPECTED_SENDER.casefold() not in addresses:
        return False

    subject = decode_mime_header(message.get("Subject", ""))
    return _normalized(SUBJECT_FRAGMENT).casefold() in subject.casefold()


def find_latest_matching_uid(
    mailbox: Any,
    *,
    now: datetime,
    max_age_hours: int,
) -> bytes | None:
    """Find the newest fresh message after locally verifying From and Subject."""
    if now.tzinfo is None:
        now = now.replace(tzinfo=timezone.utc)
    now = now.astimezone(timezone.utc)
    cutoff = now - timedelta(hours=max_age_hours)

    # IMAP SINCE only accepts a date. One extra day avoids timezone edge cases;
    # the exact hour is enforced below with the server-owned INTERNALDATE.
    search_start = (cutoff - timedelta(days=1)).date()
    criteria = (
        f'(FROM "{EXPECTED_SENDER}" SINCE {format_imap_date(search_start)})'
    )
    status, response = mailbox.uid("SEARCH", None, criteria)
    if not _status_is_ok(status):
        raise MailboxError("Не удалось выполнить поиск писем в Яндекс Почте.")

    uids: list[bytes] = []
    for block in response or []:
        if isinstance(block, bytes):
            uids.extend(block.split())

    matches: list[tuple[datetime, int, bytes]] = []
    for uid in dict.fromkeys(uids):
        status, fetch_response = mailbox.uid(
            "FETCH",
            uid,
            "(INTERNALDATE BODY.PEEK[HEADER.FIELDS (FROM SUBJECT)])",
        )
        if not _status_is_ok(status):
            raise MailboxError("Не удалось прочитать заголовки найденного письма.")

        _prefix, header_bytes = _extract_fetch_response(fetch_response)
        received_at = parse_internaldate(_internaldate_fragments(fetch_response))
        if received_at < cutoff or received_at > now + timedelta(minutes=15):
            continue

        try:
            header_message = BytesParser(policy=policy.default).parsebytes(header_bytes)
        except Exception as error:
            raise MailboxError("Не удалось разобрать заголовки найденного письма.") from error
        if not _message_matches(header_message):
            continue

        try:
            numeric_uid = int(uid)
        except (TypeError, ValueError):
            numeric_uid = 0
        matches.append((received_at, numeric_uid, uid))

    if not matches:
        return None
    return max(matches, key=lambda item: (item[0], item[1]))[2]


def fetch_full_message(mailbox: Any, uid: bytes) -> Message:
    status, response = mailbox.uid("FETCH", uid, "(BODY.PEEK[])")
    if not _status_is_ok(status):
        raise MailboxError("Не удалось скачать найденное письмо из Яндекс Почты.")
    _metadata, message_bytes = _extract_fetch_response(response)
    try:
        return BytesParser(policy=policy.default).parsebytes(message_bytes)
    except Exception as error:
        raise MailboxError("Не удалось разобрать найденное письмо.") from error


def select_xlsx_attachment(message: Message) -> AttachmentCandidate | None:
    candidates: list[AttachmentCandidate] = []
    for index, part in enumerate(message.walk()):
        if part.is_multipart():
            continue

        try:
            filename = _filename_basename(decode_mime_header(part.get_filename()))
            content_type = part.get_content_type().casefold()
            disposition = part.get_content_disposition()
        except Exception:
            continue

        normalized_name = filename.casefold()
        has_xlsx_extension = normalized_name.endswith(".xlsx")
        has_xlsx_mime_type = content_type == XLSX_MIME_TYPE
        if not (has_xlsx_extension or has_xlsx_mime_type):
            continue
        if normalized_name.startswith("~$"):
            continue

        score = 0
        if has_xlsx_extension:
            score += 100
        if "остат" in normalized_name:
            score += 60
        if "stock" in normalized_name:
            score += 40
        if has_xlsx_mime_type:
            score += 20
        if disposition == "attachment":
            score += 10
        if filename:
            score += 5
        candidates.append(AttachmentCandidate(part, filename, index, score))

    if not candidates:
        return None

    selected = max(candidates, key=lambda candidate: (candidate.score, -candidate.index))
    print(
        "Найдено XLSX-вложений: "
        f"{len(candidates)}. Выбрано: {safe_log_filename(selected.filename)}."
    )
    return selected


def attachment_payload(candidate: AttachmentCandidate) -> bytes:
    try:
        payload = candidate.part.get_payload(decode=True)
    except Exception as error:
        raise AttachmentValidationError(
            "Не удалось декодировать выбранное XLSX-вложение. Старый stock.xlsx сохранён."
        ) from error

    if not isinstance(payload, (bytes, bytearray)) or not payload:
        raise AttachmentValidationError(
            "Выбранное XLSX-вложение пустое. Старый stock.xlsx сохранён."
        )
    return bytes(payload)


def validate_xlsx(path: Path) -> None:
    """Check ZIP integrity, required XLSX parts, and workbook readability."""
    try:
        size = path.stat().st_size
    except OSError as error:
        raise AttachmentValidationError(
            "Не удалось прочитать временный XLSX-файл. Старый stock.xlsx сохранён."
        ) from error

    if size <= 0:
        raise AttachmentValidationError(
            "Получен пустой XLSX-файл. Старый stock.xlsx сохранён."
        )
    if size > MAX_ATTACHMENT_BYTES:
        raise AttachmentValidationError(
            "XLSX-вложение превышает лимит 95 МиБ для безопасной загрузки в GitHub. "
            "Старый stock.xlsx сохранён."
        )
    if not zipfile.is_zipfile(path):
        raise AttachmentValidationError(
            "Вложение не является действительным XLSX-файлом. Старый stock.xlsx сохранён."
        )

    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            member_names = {member.filename for member in members}
            missing_members = REQUIRED_XLSX_MEMBERS - member_names
            if missing_members:
                raise AttachmentValidationError(
                    "В XLSX-вложении отсутствуют обязательные части книги. "
                    "Старый stock.xlsx сохранён."
                )
            if any(member.flag_bits & 0x1 for member in members):
                raise AttachmentValidationError(
                    "Зашифрованные XLSX-вложения не поддерживаются. "
                    "Старый stock.xlsx сохранён."
                )

            uncompressed_size = sum(member.file_size for member in members)
            compressed_size = sum(member.compress_size for member in members)
            if uncompressed_size > MAX_UNCOMPRESSED_BYTES:
                raise AttachmentValidationError(
                    "Распакованный XLSX превышает безопасный лимит размера. "
                    "Старый stock.xlsx сохранён."
                )
            if (
                uncompressed_size > 50 * 1024 * 1024
                and uncompressed_size / max(compressed_size, 1)
                > MAX_SUSPICIOUS_COMPRESSION_RATIO
            ):
                raise AttachmentValidationError(
                    "XLSX-вложение имеет подозрительную степень сжатия. "
                    "Старый stock.xlsx сохранён."
                )

            damaged_member = archive.testzip()
            if damaged_member is not None:
                raise AttachmentValidationError(
                    "Контрольная сумма XLSX-вложения повреждена. "
                    "Старый stock.xlsx сохранён."
                )
    except AttachmentValidationError:
        raise
    except (
        EOFError,
        NotImplementedError,
        OSError,
        RuntimeError,
        ValueError,
        zipfile.BadZipFile,
    ) as error:
        raise AttachmentValidationError(
            "XLSX-вложение повреждено. Старый stock.xlsx сохранён."
        ) from error

    if load_workbook is None:
        raise ConfigurationError(
            "Не установлена зависимость openpyxl; XLSX не был заменён."
        )

    try:
        with path.open("rb") as workbook_stream:
            workbook = load_workbook(
                workbook_stream,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            try:
                if not workbook.sheetnames:
                    raise AttachmentValidationError(
                        "В XLSX-вложении нет листов. Старый stock.xlsx сохранён."
                    )
                # read_only parses worksheet XML lazily. Scan a bounded number of
                # rows so a forged dimension cannot consume the Actions timeout.
                rows_left = MAX_OPENPYXL_ROWS_TO_SCAN
                for worksheet in workbook.worksheets:
                    if rows_left <= 0:
                        break
                    rows_scanned = sum(
                        1
                        for _row in islice(
                            worksheet.iter_rows(values_only=True),
                            rows_left,
                        )
                    )
                    rows_left -= rows_scanned
            finally:
                workbook.close()
    except AttachmentValidationError:
        raise
    except Exception as error:
        raise AttachmentValidationError(
            "openpyxl не смог открыть XLSX-вложение: файл повреждён. "
            "Старый stock.xlsx сохранён."
        ) from error


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as error:
        raise InstallError("Не удалось вычислить хеш stock.xlsx.") from error
    return digest.hexdigest()


def _sync_directory(directory: Path) -> None:
    if not hasattr(os, "O_DIRECTORY"):
        return
    try:
        descriptor = os.open(directory, os.O_RDONLY | os.O_DIRECTORY)
        try:
            os.fsync(descriptor)
        finally:
            os.close(descriptor)
    except OSError:
        # os.replace has already completed atomically; directory fsync is an
        # additional durability measure and must not turn success into failure.
        pass


def install_payload_atomically(payload: bytes, target: Path) -> UpdateResult:
    """Stage, validate, hash, and atomically install an XLSX payload."""
    if target.exists() and not target.is_file():
        raise ConfigurationError(f"Целевой путь не является файлом: {target}")
    if target.parent.exists() and not target.parent.is_dir():
        raise ConfigurationError(f"Родительский путь не является каталогом: {target.parent}")

    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        descriptor, temporary_name = tempfile.mkstemp(
            prefix=".stock-download-",
            suffix=".xlsx",
            dir=target.parent,
        )
    except OSError as error:
        raise InstallError("Не удалось создать временный файл рядом со stock.xlsx.") from error

    temporary_path = Path(temporary_name)
    try:
        try:
            with os.fdopen(descriptor, "wb") as stream:
                stream.write(payload)
                stream.flush()
                os.fsync(stream.fileno())
        except OSError as error:
            raise InstallError("Не удалось записать временный XLSX-файл.") from error

        validate_xlsx(temporary_path)
        new_hash = sha256_file(temporary_path)

        if target.exists():
            old_hash = sha256_file(target)
            if new_hash == old_hash:
                return UpdateResult(Outcome.UNCHANGED, changed=False)
            try:
                file_mode = stat.S_IMODE(target.stat().st_mode)
            except OSError as error:
                raise InstallError("Не удалось прочитать права старого stock.xlsx.") from error
        else:
            file_mode = 0o644

        try:
            os.chmod(temporary_path, file_mode)
            os.replace(temporary_path, target)
        except OSError as error:
            raise InstallError(
                "Не удалось атомарно заменить stock.xlsx; старый файл не удалялся заранее."
            ) from error
        _sync_directory(target.parent)
        return UpdateResult(Outcome.UPDATED, changed=True)
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def process_mailbox(
    mailbox: Any,
    *,
    target: Path,
    now: datetime,
    max_age_hours: int,
) -> UpdateResult:
    uid = find_latest_matching_uid(
        mailbox,
        now=now,
        max_age_hours=max_age_hours,
    )
    if uid is None:
        return UpdateResult(Outcome.NO_MESSAGE, changed=False)

    message = fetch_full_message(mailbox, uid)
    candidate = select_xlsx_attachment(message)
    if candidate is None:
        return UpdateResult(Outcome.NO_ATTACHMENT, changed=False)

    payload = attachment_payload(candidate)
    return install_payload_atomically(payload, target)


def _read_max_age_hours() -> int:
    raw_value = os.environ.get("STOCK_MAX_EMAIL_AGE_HOURS", str(DEFAULT_MAX_AGE_HOURS))
    try:
        value = int(raw_value)
    except ValueError as error:
        raise ConfigurationError(
            "STOCK_MAX_EMAIL_AGE_HOURS должен быть целым числом."
        ) from error
    if not 1 <= value <= 24 * 30:
        raise ConfigurationError(
            "STOCK_MAX_EMAIL_AGE_HOURS должен быть от 1 до 720 часов."
        )
    return value


def run() -> UpdateResult:
    email_address = os.environ.get("YANDEX_EMAIL", "").strip()
    app_password = os.environ.get("YANDEX_APP_PASSWORD", "")
    if not email_address or not app_password:
        raise ConfigurationError(
            "Не заданы GitHub Secrets YANDEX_EMAIL и/или YANDEX_APP_PASSWORD."
        )
    if not os.environ.get("STOCK_SENDER_EMAIL", "").strip():
        raise ConfigurationError(
            "Не задан GitHub Secret STOCK_SENDER_EMAIL (e-mail поставщика)."
        )

    max_age_hours = _read_max_age_hours()
    repo_root = Path(__file__).resolve().parents[1]
    target = repo_root / TARGET_RELATIVE_PATH

    mailbox = None
    try:
        context = ssl.create_default_context()
        mailbox = imaplib.IMAP4_SSL(
            IMAP_HOST,
            IMAP_PORT,
            ssl_context=context,
            timeout=30,
        )
        mailbox.login(email_address, app_password)
        status, _response = mailbox.select(MAILBOX, readonly=True)
        if not _status_is_ok(status):
            raise MailboxError("Не удалось открыть папку INBOX в Яндекс Почте.")
        return process_mailbox(
            mailbox,
            target=target,
            now=datetime.now(timezone.utc),
            max_age_hours=max_age_hours,
        )
    except StockUpdateError:
        raise
    except imaplib.IMAP4.error as error:
        raise MailboxError(
            "Ошибка IMAP Яндекс Почты. Проверьте IMAP и GitHub Secrets."
        ) from error
    except (OSError, ssl.SSLError) as error:
        raise MailboxError(
            "Не удалось подключиться к Яндекс Почте по IMAP SSL."
        ) from error
    finally:
        if mailbox is not None:
            try:
                mailbox.logout()
            except Exception:
                pass


def write_github_outputs(result: UpdateResult) -> None:
    output_path = os.environ.get("GITHUB_OUTPUT")
    if not output_path:
        return
    try:
        with open(output_path, "a", encoding="utf-8") as output:
            output.write(f"changed={'true' if result.changed else 'false'}\n")
            output.write(f"result={result.outcome.value}\n")
    except OSError as error:
        raise ConfigurationError("Не удалось записать результат шага GitHub Actions.") from error


def _print_result(result: UpdateResult) -> None:
    messages = {
        Outcome.UPDATED: "Новый stock.xlsx проверен и атомарно установлен.",
        Outcome.UNCHANGED: "Новый XLSX совпадает с текущим stock.xlsx; commit не нужен.",
        Outcome.NO_MESSAGE: (
            "Свежих писем от поставщика с нужной темой нет; commit не нужен."
        ),
        Outcome.NO_ATTACHMENT: (
            "В самом новом подходящем письме нет XLSX-вложения; commit не нужен."
        ),
    }
    print(messages[result.outcome])


def main() -> int:
    exit_codes = {
        ConfigurationError: 2,
        MailboxError: 3,
        AttachmentValidationError: 4,
        InstallError: 5,
    }
    try:
        result = run()
        write_github_outputs(result)
    except tuple(exit_codes) as error:
        print(f"Ошибка: {error}", file=sys.stderr)
        return exit_codes[type(error)]
    except Exception as error:
        print(
            f"Непредвиденная ошибка ({type(error).__name__}); stock.xlsx не следует считать обновлённым.",
            file=sys.stderr,
        )
        return 1

    _print_result(result)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
