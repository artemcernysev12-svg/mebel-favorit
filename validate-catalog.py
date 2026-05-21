#!/usr/bin/env python3
"""
Сверяет catalog.js со свежим Авито-фидом (xlsx из личного кабинета).

Запуск:
    python3 validate-catalog.py /путь/к/Catalog_DD_MM_YYYY.xlsx

Что проверяет:
1. Все ли ID каталога есть в Авито-файле (и наоборот).
2. Не отстали ли цены.
3. Не разошлись ли названия.
4. Длину ID — все должны быть 10-значными Авито-ID.

Если найдены проблемы — выводит список и возвращает код выхода 1.
"""
import sys, re, json
from openpyxl import load_workbook

CATALOG_FILE = "catalog.js"

def load_avito(xlsx_path):
    """Извлекает Авито-ID, название, цену со всех листов Авито-файла."""
    wb = load_workbook(xlsx_path, read_only=True)
    ads = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue
        headers = [str(h or '').strip() for h in rows[3]]
        def col(name):
            for i, h in enumerate(headers):
                if h == name: return i
            return -1
        c_aid   = col('Номер объявления на Авито')
        c_name  = col('Название объявления')
        c_price = col('Цена')
        if c_aid < 0:
            continue
        for r in rows[4:]:
            if not r or c_aid >= len(r): continue
            aid = r[c_aid]
            if aid is None: continue
            s = str(aid).strip()
            if s.endswith('.0'): s = s[:-2]
            if not re.fullmatch(r'\d+', s): continue
            ads[s] = {
                'name':  str(r[c_name] or '') if c_name >= 0 and c_name < len(r) else '',
                'price': r[c_price]            if c_price >= 0 and c_price < len(r) else None,
                'sheet': sn,
            }
    wb.close()
    return ads

def load_catalog():
    with open(CATALOG_FILE, 'r', encoding='utf-8') as f:
        src = f.read()
    m = re.search(r'window\.CATALOG\s*=\s*(\[.*\])', src, re.DOTALL)
    if not m:
        raise RuntimeError("Не нашёл window.CATALOG = [...] в catalog.js")
    return json.loads(m.group(1))

def main():
    if len(sys.argv) < 2:
        print("Использование: python3 validate-catalog.py /путь/к/Catalog_xxxx.xlsx")
        sys.exit(2)
    xlsx = sys.argv[1]
    print(f"Авито-файл: {xlsx}")
    avito = load_avito(xlsx)
    catalog = load_catalog()

    print(f"\nВ catalog.js:    {len(catalog)} карточек")
    print(f"В Авито-файле:   {len(avito)} объявлений")

    cat_ids = {it['id'] for it in catalog}
    av_ids = set(avito.keys())

    only_cat = cat_ids - av_ids
    only_av  = av_ids - cat_ids

    print(f"\nID только в catalog.js (нет на Авито): {len(only_cat)}")
    if only_cat:
        for x in sorted(only_cat)[:10]:
            t = next((it.get('t','') for it in catalog if it['id']==x), '')
            print(f"  {x}  '{t[:50]}'")
        if len(only_cat) > 10:
            print(f"  ... и ещё {len(only_cat)-10}")

    print(f"\nID только в Авито (нет в catalog.js): {len(only_av)}")
    if only_av:
        for x in sorted(only_av)[:10]:
            print(f"  {x}  '{avito[x]['name'][:50]}'")

    # Длины ID в каталоге
    bad_len = [it for it in catalog
               if re.fullmatch(r'\d+', it['id']) and len(it['id']) != 10]
    print(f"\nID с длиной ≠ 10 цифр (битые): {len(bad_len)}")
    for it in bad_len[:10]:
        print(f"  {it['id']}  '{it.get('t','')[:50]}'")

    # Цены
    price_diff = 0
    for it in catalog:
        ad = avito.get(it['id'])
        if not ad or ad['price'] is None: continue
        try:
            ap = int(str(ad['price']).strip())
        except (ValueError, TypeError):
            continue
        if it.get('p', 0) != ap:
            price_diff += 1
    print(f"\nЦены расходятся: {price_diff}")

    problems = len(only_cat) + len(only_av) + len(bad_len) + price_diff
    if problems == 0:
        print("\n✅ Всё совпадает с Авито-фидом")
        return 0
    else:
        print(f"\n⚠ Найдено проблем: {problems}")
        return 1

if __name__ == "__main__":
    sys.exit(main())
